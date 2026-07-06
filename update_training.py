#!/usr/bin/env python3
"""
Update the Training Log directly inside tzr_sam_dashboard.html, then commit
and push so the published Streamlit site updates.

Usage:
    python3 update_training.py
        (uses the default training_log.xlsx on the Desktop)
    python3 update_training.py /path/to/training_log.xlsx
    python3 update_training.py ... --no-push

Expected Excel columns (first sheet): Date, Training Load, Hours, Content
"""
import sys
import re
import json
import subprocess
import datetime
from pathlib import Path

import openpyxl

REPO_DIR = Path(__file__).parent
HTML_PATH = REPO_DIR / "tzr_sam_dashboard.html"
DEFAULT_EXCEL = Path.home() / "Desktop/sam_dashboard/training_log.xlsx"

COLUMN_MAP = {
    "date": ["Date", "date"],
    "load": ["Training Load", "Load", "load"],
    "hours": ["Hours", "hours"],
    "content": ["Content", "content"],
    "why": ["Why Training That", "Why", "why"],
}


def parse_excel(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    header = [str(h).strip() if h is not None else "" for h in rows[0]]

    def col_index(candidates):
        for c in candidates:
            if c in header:
                return header.index(c)
        return None

    idx = {k: col_index(v) for k, v in COLUMN_MAP.items()}
    sessions = []
    for row in rows[1:]:
        if not row or all(v is None for v in row):
            continue

        def get(key):
            i = idx[key]
            if i is None or i >= len(row) or row[i] is None:
                return ""
            val = row[i]
            if isinstance(val, (datetime.date, datetime.datetime)):
                return val.strftime("%Y-%m-%d")
            return str(val).strip()

        date, content = get("date"), get("content")
        if not date and not content:
            continue
        sessions.append({
            "date": date,
            "load": get("load"),
            "hours": get("hours"),
            "content": content,
            "why": get("why"),
        })
    return sessions


def patch_training(html, sessions):
    pattern = re.compile(r"let TRAINING = \[.*?\];\n")
    m = pattern.search(html)
    if not m:
        raise RuntimeError("Could not find TRAINING literal in html")
    new = "let TRAINING = " + json.dumps(sessions, ensure_ascii=False) + ";\n"
    return html[: m.start()] + new + html[m.end() :]


def git(*args):
    subprocess.run(["git", *args], cwd=REPO_DIR, check=True)


def main():
    args = [a for a in sys.argv[1:] if not a.startswith("--")]
    no_push = "--no-push" in sys.argv
    excel_path = Path(args[0]) if args else DEFAULT_EXCEL

    if not excel_path.exists():
        sys.exit(f"Excel not found: {excel_path}")

    print(f"Reading excel: {excel_path}")
    sessions = parse_excel(excel_path)
    for s in sessions:
        print(f"  {s['date']}: load={s['load']} hours={s['hours']}")

    html = HTML_PATH.read_text(encoding="utf-8")
    html = patch_training(html, sessions)
    HTML_PATH.write_text(html, encoding="utf-8")
    print(f"Patched {HTML_PATH}")

    if no_push:
        print("Skipping git commit/push (--no-push).")
        return

    git("add", "tzr_sam_dashboard.html")
    diff = subprocess.run(["git", "diff", "--cached", "--quiet"], cwd=REPO_DIR)
    if diff.returncode == 0:
        print("Nothing changed, skipping commit.")
        return
    git("commit", "-m", f"Update Training Log ({len(sessions)} sessions)")
    git("push", "origin", "main")
    print("Published! Streamlit will redeploy in 1-2 minutes.")


if __name__ == "__main__":
    main()
