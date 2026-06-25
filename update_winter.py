#!/usr/bin/env python3
"""
Update the Winter season data + analysis images directly inside tzr_sam_dashboard.html,
then commit and push so the published Streamlit site updates.

Usage:
    python3 update_winter.py
        (uses the default Excel + scans the default folder for one PDF per game,
         matched by filename, e.g. "Internationale.pdf" -> the "Internationale" game)
    python3 update_winter.py /path/to/excel.xlsx /path/to/pdf_folder_or_file
    python3 update_winter.py ... --no-push

What it does:
  1. Reads the wide-format LiveTag Excel (one column per game, one row per stat)
     and rebuilds the Winter season's game list.
  2. For each game, looks for a PDF whose filename matches the game name
     (spaces/case/punctuation ignored, e.g. "Perth United" <-> "PerthUnited.pdf").
     Reads up to its first 4 pages, auto-crops each to just the chart (removing
     the yellow header band, footer and white margins), and attaches them as
     that game's analysis images — other games' images are left untouched.
  3. Patches tzr_sam_dashboard.html in this folder, commits and pushes to GitHub.
"""
import sys
import re
import json
import base64
import io
import subprocess
from pathlib import Path

import openpyxl
import fitz  # PyMuPDF
from PIL import Image
import numpy as np

REPO_DIR = Path(__file__).parent
HTML_PATH = REPO_DIR / "tzr_sam_dashboard.html"

DEFAULT_EXCEL = Path.home() / "Desktop/sam_dashboard/excel_sam.xlsx"
DEFAULT_PDF_DIR = Path.home() / "Desktop/sam_dashboard"

# Maps the WINTER schema field -> list of possible column header names in the Excel (first match wins)
FIELD_MAP = {
    "pass":        ["PASS SUCCESS", "PASS"],
    "passFail":    ["PASS FAIL"],
    "keyPass":     ["KEY PASS"],
    "shot":        ["Total SHOT", "SHOT"],
    "shot5m":      ["SHOT 5m"],
    "shot10m":     ["SHOT 10m"],
    "shotOT":      ["SHOT ON TARGET"],
    "goal":        ["GOAL"],
    "assist":      ["ASSIST"],
    "dribble":     ["DRIBBLE SUCCESS", "DRIBBLE"],
    "duel":        ["DUEL WON", "DUEL"],
    "recovery":    ["RECOVERY"],
    "lost":        ["BALL LOST", "LOSS"],
    "intercept":   ["INTERCEPTION"],
    "pressure":    ["PRESSURE"],
    "block":       ["BLOCK SHOT", "BLOCK"],
    "foulWon":     ["FOUL WON"],
    "foulCom":     ["FOUL COMMITTED"],
    "counter":     ["COUNTER ATTACK"],
    "defCover":    ["DEFENSIVE COVER"],
    "recOppPress": ["RECOVERIES OPP PRESS"],
}


def parse_excel(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    header = rows[0]
    games = [g for g in header[1:] if g is not None]
    stats = {}
    for row in rows[1:]:
        if not row or row[0] is None:
            continue
        stats[str(row[0]).strip()] = row[1:]

    winter = []
    for idx, gname in enumerate(games):
        entry = {"game": gname, "min": 0}
        for field, candidates in FIELD_MAP.items():
            val = 0
            for c in candidates:
                if c in stats and idx < len(stats[c]) and stats[c][idx] is not None:
                    val = stats[c][idx]
                    break
            entry[field] = val
        winter.append(entry)
    return winter


def normalize(name):
    return re.sub(r"[^a-z0-9]", "", name.lower())


def find_pdf_for_game(pdf_dir, game_name):
    target = normalize(game_name)
    for pdf in pdf_dir.glob("*.pdf"):
        if normalize(pdf.stem) == target:
            return pdf
    return None


def autocrop(img, pad=20, thresh=250):
    arr = np.array(img.convert("L"))
    mask = arr < thresh
    if not mask.any():
        return img
    ys, xs = mask.nonzero()
    top, bottom = max(0, ys.min() - pad), min(img.height, ys.max() + pad)
    left, right = max(0, xs.min() - pad), min(img.width, xs.max() + pad)
    return img.crop((left, top, right, bottom))


def split_chart_blocks(img, gap_thresh=250, min_gap=25):
    """LiveTag pages always stack exactly two charts (e.g. 'Goal points' above
    'Pitch points', each with its own title/diagram/legend). Splits the page
    in two at the blank horizontal band closest to the vertical middle, so
    each chart becomes its own, larger image — picking the middle gap (rather
    than just the largest) avoids cutting at title-to-diagram or trailing
    whitespace gaps near the top/bottom instead of the real chart boundary."""
    arr = np.array(img.convert("L"))
    row_has_content = (arr < gap_thresh).any(axis=1)
    gaps = []
    start = None
    for y, has in enumerate(row_has_content):
        if not has:
            if start is None:
                start = y
        else:
            if start is not None:
                gaps.append((start, y))
                start = None
    if start is not None:
        gaps.append((start, len(row_has_content)))
    h = img.height
    candidates = [g for g in gaps if g[1] - g[0] >= min_gap and g[0] > 0 and g[1] < h]
    if not candidates:
        return [autocrop(img)]
    mid = h / 2
    gap_top, gap_bottom = min(candidates, key=lambda g: abs((g[0] + g[1]) / 2 - mid))
    top_half = img.crop((0, 0, img.width, gap_top))
    bottom_half = img.crop((0, gap_bottom, img.width, h))
    return [autocrop(top_half), autocrop(bottom_half)]


def extract_pdf_images(pdf_path, zoom=3, pages=(1, 2), max_images=4):
    """Reads only the given 0-indexed pages (default: PDF pages 2 and 3 — the
    ones with the pitch/goal diagrams, skipping the page 1 comparison chart
    and any later pages), splits each page into separate chart blocks, and
    returns up to max_images base64 PNGs."""
    doc = fitz.open(pdf_path)
    blocks = []
    for p in pages:
        if p >= doc.page_count:
            continue
        pix = doc[p].get_pixmap(matrix=fitz.Matrix(zoom, zoom))
        img = Image.open(io.BytesIO(pix.tobytes("png")))
        w, h = img.size
        header_h = int(h * 62 / 1684)
        footer_h = int(h * 70 / 1684)
        sub = img.crop((0, header_h, w, h - footer_h))
        blocks.extend(split_chart_blocks(sub))
    out = []
    for block in blocks[:max_images]:
        buf = io.BytesIO()
        block.save(buf, format="PNG")
        out.append("data:image/png;base64," + base64.b64encode(buf.getvalue()).decode())
    while len(out) < 4:
        out.append("")
    return out[:4]


def patch_winter(html, winter):
    pattern = re.compile(r"let WINTER = \[.*?\];\n")
    m = pattern.search(html)
    if not m:
        raise RuntimeError("Could not find WINTER literal in html")
    new = "let WINTER = " + json.dumps(winter, ensure_ascii=False) + ";\n"
    return html[:m.start()] + new + html[m.end():]


def patch_image_default(html, varname, key, value):
    """Update one key inside the `let <varname> = JSON.parse(...) || {...};` default object."""
    start_marker = f"let {varname}"
    si = html.index(start_marker)
    # find the `) || {` that closes JSON.parse(...) and precedes the default object literal
    close_marker = ") || {"
    ci = html.index(close_marker, si)
    oi = ci + len(close_marker) - 1  # position of the opening '{'
    # the default object literal runs until its matching top-level `};`
    end = html.index("};\n", oi) + 2  # index just past the ';'
    obj_literal = html[oi:end - 1]  # from '{' up to and including the matching '}' (no trailing ';')
    obj = json.loads(obj_literal)
    obj[key] = value
    new_literal = json.dumps(obj, ensure_ascii=False) + ";"
    return html[:oi] + new_literal + html[end:]


def git(*args):
    subprocess.run(["git", *args], cwd=REPO_DIR, check=True)


CAPTION_TEMPLATES = [
    "Goal points & pitch points — shots, recoveries, interceptions",
    "Shot zones & passing lanes (pitch points)",
    "Full pitch points — passes, duels, recoveries",
    "Additional match analysis",
]


def main():
    args = [a for a in sys.argv[1:] if not a.startswith("--")]
    no_push = "--no-push" in sys.argv
    excel_path = Path(args[0]) if len(args) > 0 else DEFAULT_EXCEL
    pdf_arg = Path(args[1]) if len(args) > 1 else DEFAULT_PDF_DIR

    if not excel_path.exists():
        sys.exit(f"Excel not found: {excel_path}")
    if not pdf_arg.exists():
        sys.exit(f"PDF path not found: {pdf_arg}")
    pdf_dir = pdf_arg if pdf_arg.is_dir() else pdf_arg.parent

    print(f"Reading excel: {excel_path}")
    winter = parse_excel(excel_path)
    for g in winter:
        print(f"  {g['game']}: pass={g['pass']} shot={g['shot']} goal={g['goal']}")

    html = HTML_PATH.read_text(encoding="utf-8")
    html = patch_winter(html, winter)

    updated_games = []
    for idx, g in enumerate(winter):
        key = f"g{idx}"
        pdf_path = pdf_arg if (pdf_arg.is_file() and len(winter) == 1) else find_pdf_for_game(pdf_dir, g["game"])
        if not pdf_path:
            print(f"  (no PDF found for '{g['game']}', skipping images)")
            continue
        print(f"  Reading {pdf_path.name} -> attaching images to {g['game']} ({key})")
        images = extract_pdf_images(pdf_path)
        captions = list(CAPTION_TEMPLATES)
        html = patch_image_default(html, "wImgByGame", key, images)
        html = patch_image_default(html, "wCapsByGame", key, captions)
        updated_games.append(g["game"])

    HTML_PATH.write_text(html, encoding="utf-8")
    print(f"Patched {HTML_PATH}")

    if no_push:
        print("Skipping git commit/push (--no-push).")
        return

    git("add", "tzr_sam_dashboard.html")
    diff = subprocess.run(["git", "diff", "--cached", "--quiet"], cwd=REPO_DIR)
    if diff.returncode == 0:
        print("Nothing changed, skipping commit.")
        return
    msg = "Update Winter data" + (f" and images for {', '.join(updated_games)}" if updated_games else "")
    git("commit", "-m", msg)
    git("push", "origin", "main")
    print("Published! Streamlit will redeploy in 1-2 minutes.")


if __name__ == "__main__":
    main()
