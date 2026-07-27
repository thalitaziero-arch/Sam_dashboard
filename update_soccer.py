#!/usr/bin/env python3
"""
Update the Soccer (NPLW) season data inside tzr_sam_dashboard.html,
then commit and push.

Usage:
    python3 update_soccer.py
    python3 update_soccer.py /path/to/file.xlsx
    python3 update_soccer.py ... --no-push

Expected Excel: "Player stats S. Brady.xlsx" format exported from Wyscout.
"""
import sys
import re
import json
import subprocess
import datetime
from pathlib import Path

import openpyxl

REPO_DIR = Path(__file__).parent
HTML_PATH = REPO_DIR / "tzr_sam_dashboard.html"
DEFAULT_EXCEL = Path.home() / "Desktop/sam_dashboard/Player stats S. Brady.xlsx"


def parse_excel(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))

    def v(row, i, default=0):
        val = row[i] if i < len(row) else None
        if val is None:
            return default
        if isinstance(val, (int, float)):
            return val
        return default

    soccer = []
    for row in rows[1:]:
        if not row or row[0] is None:
            continue
        match = str(row[0]).strip()
        date = row[2]
        if isinstance(date, (datetime.date, datetime.datetime)):
            date = date.strftime("%Y-%m-%d")
        else:
            date = str(date).strip() if date else ""

        min_played = v(row, 4)
        if min_played < 1:
            continue  # skip rows with 0 minutes

        passes_total = v(row, 10)
        passes_acc   = v(row, 11)

        entry = {
            "game":         f"{match} · {date}",
            "min":          min_played,
            "goal":         v(row, 5),
            "assist":       v(row, 6),
            "shot":         v(row, 7),
            "shotOT":       v(row, 8),
            "xg":           v(row, 9, 0.0),
            "pass":         passes_acc,
            "passFail":     max(0, passes_total - passes_acc),
            "longPassAcc":  v(row, 13),
            "crossAcc":     v(row, 15),
            "dribbleTotal": v(row, 16),
            "dribble":      v(row, 17),
            "duelTotal":    v(row, 18),
            "duelWon":      v(row, 19),
            "intercept":    v(row, 20),
            "lossOwnHalf":  v(row, 21),
            "recOwnHalf":   v(row, 22),
            "recOppHalf":   v(row, 24),
            "yellowCard":   v(row, 25),
            "clearance":    v(row, 26),
            "foulCom":      v(row, 27),
            "progRun":      v(row, 28),
            "passFinal":    v(row, 29),
            "passFinalAcc": v(row, 30),
            "fwdPassAcc":   v(row, 32),
            "backPassAcc":  v(row, 34),
            "redCard":      0,
        }
        soccer.append(entry)
    return soccer


def patch_soccer(html, soccer):
    pattern = re.compile(r"let SOCCER = \[.*?\];\n")
    m = pattern.search(html)
    if not m:
        raise RuntimeError("Could not find SOCCER literal in html")
    new = "let SOCCER = " + json.dumps(soccer, ensure_ascii=False) + ";\n"
    return html[: m.start()] + new + html[m.end():]


def git(*args):
    subprocess.run(["git", *args], cwd=REPO_DIR, check=True)


def main():
    args = [a for a in sys.argv[1:] if not a.startswith("--")]
    no_push = "--no-push" in sys.argv
    excel_path = Path(args[0]) if args else DEFAULT_EXCEL

    if not excel_path.exists():
        sys.exit(f"Excel not found: {excel_path}")

    print(f"Reading: {excel_path}")
    soccer = parse_excel(excel_path)
    for g in soccer:
        print(f"  {g['game']}: min={g['min']} goal={g['goal']} pass={g['pass']}")

    html = HTML_PATH.read_text(encoding="utf-8")
    html = patch_soccer(html, soccer)
    HTML_PATH.write_text(html, encoding="utf-8")
    print(f"Patched {HTML_PATH} — {len(soccer)} matches")

    if no_push:
        print("Skipping git commit/push (--no-push).")
        return

    git("add", "tzr_sam_dashboard.html")
    diff = subprocess.run(["git", "diff", "--cached", "--quiet"], cwd=REPO_DIR)
    if diff.returncode == 0:
        print("Nothing changed, skipping commit.")
        return
    git("commit", "-m", f"Update Soccer stats ({len(soccer)} matches)")
    git("push", "origin", "main")
    print("Published! Streamlit will redeploy in 1-2 minutes.")


if __name__ == "__main__":
    main()
